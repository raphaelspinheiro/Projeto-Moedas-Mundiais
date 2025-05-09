from openpyxl import load_workbook
from openpyxl import Workbook
from transformar_tabela import criar_e_formatar_excel
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from datetime import datetime
from time import sleep

###############################################################################################################################################

# Caminhos dos arquivos e configurações de download
chromedriver_path = 'C:\\Users\\raphael.pinheiro\\chromedriver-win64'

# Configurações do Chrome
chrome_options = webdriver.ChromeOptions()
prefs = {
    "profile.default_content_settings.popups": 0,
    "download.prompt_for_download": False,
    "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
}
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("--headless")  
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920,1080")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

# Inicia o navegador invisível
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
driver.get('https://dolarhoje.com/')

###############################################################################################################################################

# Dolar:

Dolar = driver.find_element(By.ID, "nacional")
Valor_dolar = Dolar.get_attribute("value")
print("Dolar:",Valor_dolar)

# Euro:

# Caminho:

Menu_moedas = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/span")
Menu_moedas.click()

sleep(1)

Selecionar_euro = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/div/ul[1]/li[4]/a")
Selecionar_euro.click()

sleep(1)

Euro = driver.find_element(By.ID, "nacional")
Valor_euro = Euro.get_attribute("value")
print("Euro:",Valor_euro)

# Libra:

# Caminho:

Menu_moedas = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/span")
Menu_moedas.click()

sleep(1)

Selecionar_libra = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/div/ul[1]/li[18]/a")
Selecionar_libra.click()

sleep(1)

Libra = driver.find_element(By.ID, "nacional")
Valor_libra = Libra.get_attribute("value")
print("Libra:",Valor_libra)

# Peso Argentino:

# Caminho:

Menu_moedas = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/span")
Menu_moedas.click()

sleep(1)

Selecionar_Peso_Argentino = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/div/ul[1]/li[20]/a")
Selecionar_Peso_Argentino.click()

sleep(1)

Peso_Argentino = driver.find_element(By.ID, "nacional")
Valor_Peso_Argentino = Peso_Argentino.get_attribute("value")
print("Peso Argentino:",Valor_Peso_Argentino)

# Iene:

# Caminho:

Menu_moedas = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/span")
Menu_moedas.click()

sleep(1)

Selecionar_Iene = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/div/ul[1]/li[17]/a")
Selecionar_Iene.click()

sleep(2)

# Encontra o input com id='estrangeiro'
campo = driver.find_element(By.ID, "estrangeiro")

# Seleciona tudo e deleta (caso tenha conteúdo anterior)
campo.send_keys(Keys.CONTROL, 'a')  # ou Keys.COMMAND se estiver em Mac
campo.send_keys(Keys.DELETE)

# Insere o número 1
campo.send_keys("1")

sleep(2)

Iene = driver.find_element(By.ID, "nacional")
Valor_Iene = Iene.get_attribute("value")
print("Iene:",Valor_Iene)

# Franco Suíço:

# Caminho:

Menu_moedas = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/span")
Menu_moedas.click()

sleep(1)

Selecionar_Franco = driver.find_element(By.XPATH, "/html/body/div[5]/ul/li/ol/li[2]/div/ul[1]/li[16]/a")
Selecionar_Franco.click()

sleep(2)

Franco = driver.find_element(By.ID, "nacional")
Valor_Franco = Franco.get_attribute("value")
print("Franco Suiço",Valor_Franco)

driver.quit()

data_hoje = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

arquivo_excel = 'moedas.xlsx'
try:
    workbook = load_workbook(arquivo_excel)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Local', 'Valor','Data' ])

# Adicionar a informação sobre o local, a data e o valor ao arquivo Excel
sheet.append(['Estados Unidos', 'Dolar', Valor_dolar, data_hoje])
sheet.append(['Portugal', 'Euro', Valor_euro, data_hoje])
sheet.append(['Argentina', 'Peso Argentino', Valor_Peso_Argentino, data_hoje])
sheet.append(['Reino Unido', 'Libra',Valor_libra, data_hoje])
sheet.append(['Japão','Iene',Valor_Iene, data_hoje])
sheet.append(['Suíça', 'Franco Suiço', Valor_Franco, data_hoje])

# Salvar o arquivo Excel
workbook.save(arquivo_excel)

# Puxando a função de criar tabela
nome_arquivo_entrada = arquivo_excel
novo_nome_arquivo = 'tabela_formatada1.xlsx'
criar_e_formatar_excel(nome_arquivo_entrada, novo_nome_arquivo)



############################################################################################################